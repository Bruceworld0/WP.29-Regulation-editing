from googletrans import Translator
from openpyxl import Workbook
from openpyxl import load_workbook
from docx import Document

translator = Translator()
write_wb = Workbook()

def createsheet(a):
    for i in a:
        write_wb.create_sheet(i)

def store(a, b):
    load_ws = write_wb[a]
    doc = Document(b)
    paragraphs = [paragraph.text for paragraph in doc.paragraphs]
    for i, para in enumerate(paragraphs):
        load_ws.cell(i+1, 1, para)
        
        try:
            output = translator.translate(para, src='en', dest='ko')
            load_ws.cell(i+1, 2, output.text)
        except :
            pass

sheets = ["0000", "0001", "0002", "0003", "0004", "0005"]

documents = ["R148e.docx", "R148am1e.docx", "R148am2e.docx", "R148am3e.docx", "R148am4e.docx", "R148am5e.docx", "R148am6e.docx"]

createsheet(sheets)

for a, b in zip(sheets, documents):
    store(a,b)

write_wb.save("UNR148_en_ko.xlsx")
