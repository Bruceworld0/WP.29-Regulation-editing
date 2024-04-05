from googletrans import Translator
from openpyxl import Workbook
from openpyxl import load_workbook
from docx import Document

translator = Translator()
write_wb = Workbook()

def createsheet(a):
    for i in a:
        write_wb.create_sheet(i)

def store(a, b):
    load_ws = write_wb[a]
    doc = Document(b)
    paragraphs = [paragraph.text for paragraph in doc.paragraphs]
    for i, para in enumerate(paragraphs):
        load_ws.cell(i+1, 1, para)
        
        try:
            output = translator.translate(para, src='en', dest='ko')
            load_ws.cell(i+1, 2, output.text)
        except :
            pass

sheets = ["0000", "0001", "0002", "0003", "0004", "0005", "0006", "0101"]

documents = ["R150e.docx", "R150am1e.docx", "R150am2e.docx", "R150am3e.docx", "R150am4e.docx", "R150am5e.docx", "R150am6e.docx",  \
              "R150r1am1e.docx"]

createsheet(sheets)

for a, b in zip(sheets, documents):
    store(a,b)

write_wb.save("UNR150_en_ko.xlsx")
