from googletrans import Translator
from openpyxl import Workbook
from openpyxl import load_workbook
from docx import Document

translator = Translator()
write_wb = Workbook()

def createsheet(a):
    for i in a:
        write_wb.create_sheet(i)

def store(a, b):
    load_ws = write_wb[a]
    doc = Document(b)
    paragraphs = [paragraph.text for paragraph in doc.paragraphs]
    for i, para in enumerate(paragraphs):
        load_ws.cell(i+1, 1, para)
        
        try:
            output = translator.translate(para, src='en', dest='ko')
            load_ws.cell(i+1, 2, output.text)
        except :
            pass

sheets = ["0000", "0001", "0002", "0003", "0004", "0005", "0006", "0007", "0008", "0101"]

documents = ["R149e.docx", "R149am1e.docx", "R149am2e.docx", "R149am3e.docx", "R149am4e.docx", "R149am5e.docx", "R149am6e.docx", "R149am7e.docx", "R149am8e.docx", \
              "R149r1am1e.docx"]

createsheet(sheets)

for a, b in zip(sheets, documents):
    store(a,b)

write_wb.save("UNR149_en_ko.xlsx")
